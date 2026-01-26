# Created on September 24th of 2024
# By: Susana Constenla

# Code to clean the raw MUAC data shared by the NDMA

# This part of the code just reads the MUAC data collected in the sentinel sites, /// 
# in excel format, and generates a pandas/python dataset with the desired columns. ///
# This is the initialization of the pipeline, and the way we will start updating ///
# the results with newly collected data every month.

# Datasets:
#	-IN: excel file with raw MUAC data (at the child level). ///
#              All in out sheet (named MUAC), /// 
#              with all wards and dates in that sheet.
#	-OUT: pickle file with the raw MUAC data in Python (pandas) format.
#==============================================================
import pandas as pd
import os
import geopandas as gpd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.patches as mpatches
import matplotlib.patheffects as path_effects
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.colors as mcolors
from scipy import stats
from matplotlib.colors import ListedColormap, BoundaryNorm

import re, unicodedata

#==============================================================
# Set input and output data path

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
os.chdir(PARENT_DIR)
 
RAW = os.path.dirname(SCRIPT_DIR)
SHAPE = os.path.join(PARENT_DIR, "shapefiles")

output_dir = "intermediary_datasets"
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
OUTPUT = os.path.join(PARENT_DIR, "intermediary_datasets")

#==============================================================
# List of 23 ASAL counties

counties = ['Baringo', 'Embu', 'Garissa', 'Isiolo', 'Kajiado', 'Kilifi', 'Kitui', 'Kwale', 
            'Laikipia', 'Lamu', 'Makueni', 'Mandera', 'Marsabit', 'Meru', 'Narok', 'Nyeri', 
            'Samburu', 'Taita_taveta', 'Tana_river', 'Tharaka_nithi', 'Turkana', 'Wajir', 'West_pokot']

cols = ['QID', 'County', 'SubCounty', 'Ward', 'LivelihoodZone', 'Month', 'Year',
        'HouseholdCode', 'ChildName', 'Gender', 'MUAC', 'MUAC_Color',
        'AgeInMonths', 'LiveInHousehold', 'SufferedIllnesses', 'InterviewDate',
        'DivisionID', 'CountyID', 'SiteID', 'LivelihoodZoneID']

#--------------------------------------------------------------------------
# Loop over counties to import and save relevant Excel data (2021-onwards)
#--------------------------------------------------------------------------
'''all_counties = pd.concat(
    [
        pd.read_excel(os.path.join(RAW, f"{county}.xlsx"), sheet_name="MUAC DEWS")[cols]
        .loc[lambda df: pd.to_numeric(df['Year'], errors='coerce') >= 2021]
        for county in counties
        if print(f"Processing {county}...") is None or True  
    ],
    ignore_index=True
)
'''
#==============================================================
new_MUAC_data_file_name = "MUAC_Data.xlsx"

preview = pd.read_excel(os.path.join(RAW
                        , new_MUAC_data_file_name), sheet_name="MUAC",
                          header=None, nrows=10)

header_row_index = None
for i, row in preview.iterrows():
    non_null = row.dropna()
    if len(non_null) >= 2 and non_null.map(type).eq(str).mean() > 0.8:
        header_row_index = i
        break

if header_row_index is None:
    raise ValueError("Could not detect header row")

muac = pd.read_excel(os.path.join(RAW
                        , new_MUAC_data_file_name), sheet_name="MUAC", 
                        header=header_row_index)

# Adding numeric month variable
muac['InterviewDate'] = pd.to_datetime(muac['InterviewDate'], errors='coerce')
muac['month_num'] = muac['InterviewDate'].dt.month
print(muac[['InterviewDate', 'month_num']].head())


final_cols = ['QID', 'County', 'SubCounty', 'Ward', 'LivelihoodZone', 'Month', 'month_num', 'Year',
        'HouseholdCode', 'ChildName', 'Gender', 'MUAC', 'MUAC_Color',
        'AgeInMonths', 'LiveInHousehold', 'SufferedIllnesses', 'InterviewDate',
        'DivisionID', 'CountyID', 'SiteID', 'LivelihoodZoneID']

final_df = muac[final_cols]

# Apply proper title case and clean text columns
import re, unicodedata

def clean_name(s):
    if pd.isna(s):
        return pd.NA
    s = str(s)

    # Unicode normalization + remove zero-widths/BOM
    s = unicodedata.normalize('NFKC', s)
    s = re.sub(r'[\u200B-\u200D\uFEFF]', '', s)

    # Standardize spaces/separators
    s = s.replace('\u00A0', ' ')
    s = re.sub(r'[_]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'\s*/\s*', '/', s)
    s = re.sub(r'\s*-\s*', '-', s)

    # Trim stray punctuation at ends
    s = s.strip(" '\".,;:()[]{}")

    # Title case
    s = s.lower().title()
    return s

for col in ['Ward', 'SubCounty', 'County']:
    if col in final_df.columns:
        final_df[col] = final_df[col].apply(clean_name)


#==============================================================
# Fix discrepancies in ward nomenclature between the ward shapefile
# (generated by FAO) and NDMA's dataset 
#==============================================================

wards = gpd.read_file(os.path.join(SHAPE, 'Kenya_wards_NDMA.shp'))

wards.plot(color='lightblue', edgecolor='black')
plt.title("Kenya Wards")
#plt.show()
#plt.close()

common_values = muac['Ward'].isin(wards['Ward'])

common_muac = muac[common_values]
unmatched_muac = muac[~common_values]

# Use geocode fix to match shapefile with gov. data
from _ward_codebook_match import replace_wards_and_counties 

muac = replace_wards_and_counties(muac)
common_values = muac['Ward'].isin(wards['Ward'])
unmatched_muac = muac[~common_values]
unmatched_muac[['Ward', 'County', 'SubCounty']].drop_duplicates()


#==============================================================
# Quality control on number of child measurements by ward and month
# Data cleaning for unfeasible values and generating wasting prevalence
# at the ward level
#==============================================================

directory = "data_quality_checks"

if not os.path.exists(directory):
    os.makedirs(directory)

muac['Year'] = pd.to_numeric(muac['Year'], errors='coerce')
muac['Year'] = muac['Year'].astype('Int64') 

counties = gpd.read_file(os.path.join(SHAPE, 
                                "ken_admbnda_adm1_iebc_20191031.shp"))

muac_summary = muac[['MUAC', 'Ward']].groupby('Ward', as_index=False)['MUAC'].count()
valid_wards = wards[wards['Ward'].isin(muac_summary['Ward'])]
county_names = [
    'Baringo', 'Embu', 'Garissa', 'Isiolo', 'Kajiado', 'Kilifi', 'Kitui', 'Kwale', 
    'Laikipia', 'Lamu', 'Makueni', 'Mandera', 'Marsabit', 'Meru', 'Narok', 'Nyeri', 
    'Samburu', 'Taita Taveta', 'Tana River', 'Tharaka-Nithi', 'Turkana', 'Wajir', 'West Pokot'
]

valid_counties = counties[counties['ADM1_EN'].isin(county_names)]

#-------------------------------------------
# Plotting wards in dataset 
#-------------------------------------------
fig, ax = plt.subplots(figsize=(14, 14))
valid_wards.plot(ax=ax, color="lightblue", 
                 edgecolor="gray", linewidth=1.5, label="MUAC Wards")

valid_counties.plot(ax=ax, color="lightgray", edgecolor="black", 
                    linewidth=2, alpha=0.5, label="County Borders")
for idx, row in valid_counties.iterrows():
    centroid = row.geometry.centroid
    ax.text(centroid.x, centroid.y, row['ADM1_EN'], 
            fontsize=12, color='darkred', ha='center',  weight='bold')
ax.legend(loc="upper left")
ax.set_axis_off()
plt.tight_layout()
#plt.show()
#plt.close()
#-------------------------------------------

last_year = muac['Year'].max()
last_month = muac[muac['Year'] == last_year]['month_num'].max()
print(f"Last year-month: {last_year}-{last_month}")

last_month_observations = muac[
    (muac['Year'] == last_year) & (muac['month_num'] == last_month)
].groupby('Ward')['MUAC'].count().reset_index(name='num_observations')

ward_county = muac[['Ward', 'County']].drop_duplicates(subset='Ward')
last_month_observations = last_month_observations.merge(
    ward_county, 
    on='Ward',
    how='left'
)

last_month_observations = last_month_observations.sort_values(
    'num_observations', ascending=False)

#-------------------------------------------------------------------
# Color coded map of number of observations in the last collected month
#--------------------------------------------------------------------
bins = [0, 50, 100, 150, np.inf]
labels = ['0–50', '50–100', '100–150', '>150']

category_colors = {
    '0–50': '#fee8c8',
    '50–100': '#fdbb84',
    '100–150': '#fc8d59',
    '>150': '#b30000'
}

last_month_observations['obs_category'] = pd.cut(
    last_month_observations['num_observations'],
    bins=bins,
    labels=labels,
    right=False
)

last_month_observations['color'] = last_month_observations['obs_category'].map(category_colors)
wards_plot = wards.merge(
    last_month_observations[['Ward', 'num_observations', 'obs_category', 'color']],
    on='Ward',
    how='left'
)

fig, ax = plt.subplots(1, 1, figsize=(12, 10))
wards_plot.plot(
    column='obs_category',
    ax=ax,
    cmap=mcolors.ListedColormap([category_colors[label] for label in labels]),
    legend=True,
    missing_kwds={"color": "lightgrey", "label": "No data"},
    edgecolor='black',
    linewidth=0.2
)

patches = [mpatches.Patch(color=category_colors[label], label=label) for label in labels]
ax.legend(handles=patches, title='Total MUAC Observations')

valid_counties.plot(ax=ax, color="lightgray", edgecolor="black",
                    linewidth=4, alpha=0.2)
for idx, row in valid_counties.iterrows():
    centroid = row.geometry.centroid
    text = ax.text(
        centroid.x, centroid.y, row['ADM1_EN'],
        fontsize=9, color='black', ha='center', weight='bold'
    )
    text.set_path_effects([
        path_effects.Stroke(linewidth=3, foreground='white'),
        path_effects.Normal()
    ])

ax.set_title(f'Total MUAC Observations per Ward\n({last_year}-{last_month:02d})')
ax.set_axis_off()
plt.tight_layout()
fig.savefig(
    f"data_quality_checks/TOTAL_muac_obs_map_{last_year}_{last_month:02d}.png",
    dpi=300,
    bbox_inches='tight',
    facecolor='white'
)
#plt.show()

#--------------------------------------
# Export ward observation numbers to Excel
#--------------------------------------
export_df = last_month_observations[['County', 'Ward', 
                                     'num_observations', 'obs_category', 'color']]
category_order = ['0–50', '50–100', '100–150', '>150']
export_df['obs_category'] = pd.Categorical(
    export_df['obs_category'],
    categories=category_order,
    ordered=True
)
export_df = export_df.sort_values(
    by=['obs_category', 'County'],
    ascending=[False, True]
)
export_path = f"data_quality_checks/ward_observation_categories_{last_year}_{last_month}.xlsx"
export_df.to_excel(export_path, index=False)
wb = load_workbook(export_path)
ws = wb.active
num_obs_col_idx = export_df.columns.get_loc('num_observations') + 1
for i, (_, row) in enumerate(export_df.iterrows(), start=2):
    hex_color = row['color'].replace('#', '')
    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
    ws.cell(row=i, column=num_obs_col_idx).fill = fill

color_col_idx = export_df.columns.get_loc('color') + 1
ws.delete_cols(color_col_idx)
wb.save(export_path)

#=============================================================================
# Cleaning MUAC data
#=============================================================================

muac['MUAC'] = pd.to_numeric(muac['MUAC'], errors='coerce')
muac['MUAC'] = muac['MUAC'].astype('Int64')

df_clean = muac.dropna(subset=['MUAC'])

# Keep children between 6 months and 5 years of age
df_clean['AgeInMonths'] = pd.to_numeric(muac['AgeInMonths'], errors='coerce')
df_clean['AgeInMonths'] = df_clean['AgeInMonths'].astype('Int64') 

df_clean.AgeInMonths.describe()
df_clean[df_clean['AgeInMonths'] <0]
df_clean.loc[df_clean['AgeInMonths'] < 0, 'AgeInMonths'] *= -1

df_clean = df_clean[(df_clean['AgeInMonths'] >= 6) & (df_clean['AgeInMonths'] < 61)]

# Drop unrealistic values
df_clean = df_clean[(df_clean['MUAC'] >= 80) & (df_clean['MUAC'] <= 250)]

# Observations per ward and month, with no missing MUAC values
observations = df_clean.groupby(
    ['Year', 'month_num', 'Ward'])['MUAC'].count().reset_index(name='num_observations')

stats = observations.groupby(['Year', 'month_num']).agg(
    avg_observations=('num_observations', 'mean'),
    std_observations=('num_observations', 'std')
).reset_index()

print(stats)

df_last_month = df_clean[
    (df_clean['Year'] == last_year) &
    (df_clean['month_num'] == last_month)
]

muac_obs = df_last_month.groupby('Ward')['MUAC'] \
    .apply(lambda x: x.notnull().sum()) \
    .reset_index(name='MUAC_obs')

print(muac_obs.head())

wards = wards.drop(columns=['MUAC_obs'], errors='ignore')
wards = wards.merge(muac_obs, on='Ward', how='left')
wards['MUAC_obs_category'] = pd.cut(wards['MUAC_obs'], 
                                    bins=bins, 
                                    labels=labels, 
                                    right=False)

fig, ax = plt.subplots(1, 1, figsize=(12, 10))

wards.plot(
    column='MUAC_obs_category',
    ax=ax,
    legend=True,
    cmap=mcolors.ListedColormap(list(category_colors.values())),
    missing_kwds={"color": "lightgrey", "label": "No data"},
    edgecolor='black',
    linewidth=0.2
)

ax.set_title(f'Valid MUAC Observations per Ward\n({last_year}-{last_month:02d})')
ax.set_axis_off()

plt.tight_layout()
patches = [mpatches.Patch(color=color, label=cat) 
           for cat, color in category_colors.items()]
valid_counties.plot(ax=ax, color="lightgray", edgecolor="black", 
                    linewidth=4, alpha=0.2, label="County Borders")

for idx, row in valid_counties.iterrows():
    centroid = row.geometry.centroid
    text = ax.text(
        centroid.x, centroid.y, row['ADM1_EN'],
        fontsize=9, color='black', ha='center', weight='bold'
    )
    text.set_path_effects([
        path_effects.Stroke(linewidth=3, foreground='white'),
        path_effects.Normal()
    ])

fig.savefig(
    f"data_quality_checks/VALID_muac_obs_map_{last_year}_{last_month:02d}.png",
    dpi=300,
    bbox_inches='tight',
    facecolor='white'
)

#plt.show()

#=================================================================
# Plotting total number of wasted children observations, per ward
#=================================================================

all_wards_in_data = df_clean[['Ward']].drop_duplicates()

df_last_month = df_clean[
    (df_clean['Year'] == last_year) &
    (df_clean['month_num'] == last_month)
]

all_wards_in_data = df_last_month[['Ward']].drop_duplicates()
muac_below_125_counts = (
    df_last_month[df_last_month['MUAC'] < 125]
    .groupby('Ward')['MUAC'].count()
    .reindex(all_wards_in_data['Ward'])
    .fillna(0)
    .reset_index()
)
muac_below_125_counts.rename(columns={'MUAC': 'MUAC_below_125_count'}, inplace=True)
wards = wards.drop(columns=['MUAC_below_125_count'], errors='ignore')
wards = wards.merge(muac_below_125_counts, on='Ward', how='left')
wards['MUAC_below_125_count'] = wards['MUAC_below_125_count'].where(
    wards['MUAC_below_125_count'].notna(), -2)
wards['MUAC_below_125_count'] = wards['MUAC_below_125_count'].replace(0, -1)

vmin = 1
vmax = wards['MUAC_below_125_count'].max()
if vmax >= vmin:
    bins = np.linspace(vmin, vmax, 20)
else:
    bins = [vmin]

colors = ['lightgrey', 'green'] + list(plt.cm.Reds(np.linspace(0, 1, len(bins))))
cmap = ListedColormap(colors)
norm = BoundaryNorm([-2, -1, 0] + list(bins), cmap.N)

fig, ax = plt.subplots(1, 1, figsize=(12, 10))

plot = wards.plot(
    column='MUAC_below_125_count',
    ax=ax,
    cmap=cmap,
    norm=norm,
    legend=True,
    legend_kwds={
        'label': "MUAC < 125 Counts",
        'orientation': "horizontal",
        'shrink': 0.7,
        'pad': 0.02
    }
)

valid_counties.plot(
    ax=ax,
    color='none',
    edgecolor='black',
    linewidth=0.8,
    label='County Boundaries'
)

ax.set_title(
    f'MUAC Observations Below 125mm per Ward ({last_year}-{last_month:02d})',
    fontsize=14,
    weight='bold'
)
ax.set_axis_off()
ax.legend(loc='upper left')

plt.tight_layout()

fig.savefig(
    f"data_quality_checks/wasted_obs_map_{last_year}_{last_month:02d}.png",
    dpi=300,
    bbox_inches='tight',
    facecolor='white'
)

#plt.show()

# Generating wasting prevalence 
df_clean['wasting_risk'] = np.where(df_clean['MUAC'] < 135, 1, 0)
df_clean['wasting'] = np.where(df_clean['MUAC'] < 125, 1, 0)
df_clean['wasting_severe'] = np.where(df_clean['MUAC'] < 115, 1, 0)

df_clean['time'] = df_clean['Year'] + (df_clean['month_num'] - 1) / 12

wasting_count = (
    df_clean.groupby(['Ward', 'Year','month_num'], as_index=False)['wasting'].sum()
    .rename(columns={'wasting': 'wasting_count'})
)

wasting_risk_count = (
    df_clean.groupby(['Ward', 'Year','month_num'], as_index=False)['wasting_risk'].sum()
    .rename(columns={'wasting_risk': 'wasting_risk_count'})
)

# Saving results at ward/month level
wasting_prevalence = df_clean.groupby(['Ward', 'Year', 'month_num'])['wasting'].mean().reset_index()
wasting_risk = df_clean.groupby(['Ward', 'Year', 'month_num'])['wasting_risk'].mean().reset_index()

columns_to_keep = ['County', 'SubCounty', 
                   'Ward', 'LivelihoodZone', 'Month', 'month_num', 'Year','time']

df_unique = df_clean[columns_to_keep].drop_duplicates()
observations_per_month = df_clean.groupby(['Ward', 'Year', 'month_num'])['MUAC'].count().reset_index()
observations_per_month.rename(columns={'MUAC': 'obs_per_month'}, inplace=True)

final_df = pd.merge(df_unique, wasting_prevalence, on=['Ward', 'Year', 'month_num'], how='left')
final_df = pd.merge(final_df, wasting_risk, on=['Ward', 'Year', 'month_num'], how='left')
final_df = pd.merge(final_df, observations_per_month, on=['Ward', 'Year', 'month_num'], how='left')
final_df = pd.merge(final_df, wasting_count, on=['Ward', 'Year', 'month_num'], how='left')
final_df = pd.merge(final_df, wasting_risk_count, on=['Ward', 'Year', 'month_num'], how='left')

final_df.rename(columns={'avg_wasting_prevalence': 'wasting'}, inplace=True)

final_df.head()

min_date = muac['InterviewDate'].min()
max_date = muac['InterviewDate'].max()

start_label = min_date.strftime('%b_%Y')
end_label = max_date.strftime('%b_%Y')

if start_label == end_label:
    # filename = f"Kenya_NDMA_MUAC_23_counties_{start_label}_WARD_LEVEL.pkl"
    filename = f"Kenya_NDMA_MUAC_23_counties.pkl"    
else:
    # filename = f"Kenya_NDMA_MUAC_23_counties_{start_label}_{end_label}_WARD_LEVEL.pkl"
    filename = f"Kenya_NDMA_MUAC_23_counties.pkl"

output_path = os.path.join(OUTPUT, filename)
final_df.to_pickle(output_path)

print(f"Saved to: {output_path}")